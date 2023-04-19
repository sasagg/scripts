# To add a new cell, type '# %%'
# To add a new markdown cell, type '# %% [markdown]'
# %%
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl
import numpy as np
import os

keyword = "エンジニア　採用"
hiduke = "20230118"
#saiyoudo_word = ""
#件数
number = 10
block_list = []

def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"].value = "会社名"
    ws["B1"].value = "会社URL"
    ws["C1"].value = "サービス名"
    ws["D1"].value = "サービスURL"
    ws["E1"].value = "業界"
    ws["F1"].value = "メールアドレス"
    ws["G1"].value = "採用ページ"
    ws["H1"].value = "採用積極度"
    ws["I1"].value = "記事の作成時間"
    ws["J1"].value = "検索ワード"
    ws["K1"].value = "PR_TIMES_URL"

    #chrome driverを解凍し、URLを張る
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    driver_path = ".\chromedriver_win32\chromedriver"
    URL = 'https://prtimes.jp/main/action.php?run=html&page=searchkey&search_word={}'.format(keyword)

    options = webdriver.ChromeOptions()
    #options.add_argument("–headless")
    #options.add_argument("–no-sandbox")
    #options.add_argument("–disable-dev-shm-usage")
    driver = webdriver.Chrome(driver_path,options=options)

    #指定したURLに遷移
    driver.get(URL)
    #カレントページのURLを取得
    cur_url = driver.current_url

    flg = False
    while flg == True:
        #ページ上の日付を取得
        date_list = [elem_h3.get_attribute("datetime").split("T")[0].replace("-","") for elem_h3 in driver.find_elements_by_xpath('//time') if not elem_h3.get_attribute("datetime")==None]
        if hiduke in date_list:
            flg = True
        else:
            #クリックしたい要素を取得
            a_item = driver.find_element(By.CSS_SELECTOR, ("div[class='list-article__more js-list-article-more-wrap']"))
            a_item = a_item.find_element(By.TAG_NAME, ("a"))
            a_item.click()
    count = 2
    for elem_h3 in driver.find_elements_by_xpath('//article/a'):
        try:
            PR_TIMES_URL = elem_h3.get_attribute("href")
        except:
            PR_TIMES_URL = "エラー"
        driver2 = webdriver.Chrome(driver_path,options=options)
        driver2.get(PR_TIMES_URL)
        for elem_h4 in driver2.find_elements_by_xpath("//div/time"):
            try:
                TIME = elem_h4.get_attribute("datetime").split("T")[0].replace("-","/")
            except:
                TIME = "エラー"
            break
        try:
            COMPANY_NAME = driver2.find_element_by_class_name("company-name").text
        except:
            COMPANY_NAME = "エラー"
        try:
            GYOKAI = driver2.find_elements_by_class_name("body-information")[1].text
        except:
            GYOKAI = "エラー"
        for elem_h4 in driver2.find_elements_by_xpath("//span/a"):
            try:
                COMPANY_URL = elem_h4.get_attribute("href")
            except:
                COMPANY_URL = "エラー"
            break
        driver2.get(COMPANY_URL)
        try:
            SAIYO_URL = driver2.find_element(By.XPATH, '//a[text()="採用情報"]').get_attribute("href")
        except:
            SAIYO_URL = "エラー"
        driver2.close()
        ws["A"+str(count)].value = COMPANY_NAME
        ws["B"+str(count)].value = COMPANY_URL
        #ws["C"+str(count)].value = SERVICE_NAME
        #ws["D"+str(count)].value = SERVICE_URL
        ws["E"+str(count)].value = GYOKAI
        #ws["F"+str(count)].value = MAIL
        ws["G"+str(count)].value = SAIYO_URL
        #ws["H"+str(count)].value = SAIYO_SEKYOKUDO
        ws["I"+str(count)].value = TIME
        ws["J"+str(count)].value = keyword
        ws["K"+str(count)].value = PR_TIMES_URL
        count +=1
        if count > number:
            break
    driver.close()    
    wb.save('./list.xlsx')

if __name__ == "__main__":
    main()


